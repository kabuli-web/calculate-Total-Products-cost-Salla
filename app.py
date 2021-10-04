import openpyxl
import os
from pprint import pprint
path = os.getcwd()+ r"\file.xlsx"
print(path)
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
total_rows = sheet_obj.max_row
product_name_column_no = 3
product_type_column_number = 2
product_quantity_column_number = 8
product_cost_column_number = 12
product_no_column_number = 1

class Product():
    def __init__(self,name,type,cost,quantity,no):
        self.type = type
        self.cost = cost
        self.quantity = quantity
        self.name = name
        self.no = no
total = 0
products = []
temp_products = []
deleted_products = []
for i in range(3,total_rows):
    
    name = sheet_obj.cell(i,product_name_column_no).value
    type = sheet_obj.cell(i,product_type_column_number).value
    quantity = sheet_obj.cell(i,product_quantity_column_number).value
    cost = sheet_obj.cell(i,product_cost_column_number).value
    no = sheet_obj.cell(i,product_no_column_number).value

    current_product = Product(name,type,cost,quantity,no)
    
    if(len(temp_products)==0):
        temp_products.append(current_product)
        continue
    last_list_prod = temp_products[len(temp_products)-1]
    
    if(last_list_prod.name==current_product.name and last_list_prod.type=="منتج"):
        temp_products.remove(last_list_prod)
        deleted_products.append(last_list_prod)
        temp_products.append(current_product)
        continue
    if(last_list_prod.name==current_product.name and last_list_prod.type!="منتج"):
        temp_products.append(current_product)
        continue
    if(last_list_prod.name != current_product.name):
        for p in temp_products:
            products.append(p)
        temp_products.clear()
        temp_products.append(current_product)
        continue

if(len(temp_products)>0):
    for p in temp_products:
            products.append(p)


for prod in products:
    if(isinstance(prod.cost,int) and isinstance(prod.quantity,int)):
        total = total + (prod.cost * prod.quantity)
    # if(isinstance(prod.cost,int)==False ):
    #     print(prod.no)
   

print('Number of products: ' + str(len(products)))
print("Total cost of Porducts "+ str(total))

